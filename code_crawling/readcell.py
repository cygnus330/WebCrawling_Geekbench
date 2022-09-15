from openpyxl import load_workbook

'''
import pandas as pd

def readexcel():
    cellread = pd.read_excel("CPUlist.xlsx")
    return cellread
'''
def readCPUlist():
    wb = load_workbook("CPUlist.xlsx") #CPUlist.xlsx 파일에서 wb를 불러옴
    ws = wb.active #활성화된 시트
    return ws

def makeCPUlist():
    CPU_RAWlist = readCPUlist()
    CPU_list = []
    for gen in range(2, 14):
        mylist = []
        mylist.append(gen-1)
        mylist.append(CPU_RAWlist.cell(row=gen, column=3).value)
        mylist.append(CPU_RAWlist.cell(row=gen, column=4).value)
        mylist.append(CPU_RAWlist.cell(row=gen, column=5).value)
        mylist.append(CPU_RAWlist.cell(row=gen, column=6).value)
        CPU_list.append(mylist)
    return CPU_list

def makeScorelist():
    Score_list = []
    for gen in range(2, 14):
        Score_list.append([gen-1, [0.0, 0.0], [0.0, 0.0], [0.0, 0.0], [0.0, 0.0]])
    return Score_list

def makeTestlist():
    test_list = []
    for gen in range(2,14):
        test_list.append([gen-1, 0, 0, 0, 0])
    return test_list